"""
=============================================================
  PNL DIARIO — Punto Casa de Bolsa
  App Streamlit: sube el COEBMV101 y descarga el PNL listo
=============================================================
"""

import io
import json
import datetime
import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ─────────────────────────────────────────────────────────────────────────────
#  CONFIGURACIÓN DE PÁGINA
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="PNL Diario — Punto",
    page_icon="📊",
    layout="wide",
)

st.markdown("""
<style>
    .main-header {
        background: linear-gradient(135deg, #1A1A2E 0%, #533483 100%);
        padding: 1.4rem 1.8rem; border-radius: 12px; margin-bottom: 1.2rem;
    }
    .main-header h1 { color: #FFFFFF; margin: 0; font-size: 1.6rem; font-weight: 700; }
    .main-header p  { color: #C9BBE0; margin: 0.2rem 0 0 0; font-size: 0.9rem; }
    [data-testid="stMetricValue"] { font-size: 1.5rem; }
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div class="main-header">
    <h1>📊 Reporte PNL Diario</h1>
    <p>Punto Casa de Bolsa · Mesa de Capitales · Cuenta 9991</p>
</div>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
#  LÓGICA DE PARSEO
# ─────────────────────────────────────────────────────────────────────────────
def parsear_coebmv(file_bytes):
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    filas = list(ws.iter_rows(values_only=True))

    idx_header = None
    for i, fila in enumerate(filas):
        if any(isinstance(c, str) and 'CONTRATO' in c.upper() for c in fila):
            idx_header = i
            break
    if idx_header is None:
        raise ValueError("No encontré la fila de encabezados ('CONTRATO') en el archivo.")

    header = filas[idx_header]
    data   = filas[idx_header + 1:]
    df = pd.DataFrame(data, columns=header)
    df.columns = [str(c).strip() if c is not None else '' for c in df.columns]

    mapa = {}
    for col in df.columns:
        cu = col.upper()
        if 'CONTRATO'   in cu: mapa[col] = 'CONTRATO'
        elif 'ORDEN'    in cu: mapa[col] = 'NO_ORDEN'
        elif 'FECHA'    in cu: mapa[col] = 'FECHA'
        elif 'COMPRA'   in cu: mapa[col] = 'TIPO'
        elif 'EMISORA'  in cu: mapa[col] = 'EMISORA'
        elif 'TITULO'   in cu: mapa[col] = 'TITULOS'
        elif 'PRECIO'   in cu: mapa[col] = 'PRECIO'
        elif 'IMPORTE'  in cu: mapa[col] = 'IMPORTE_BRUTO'
        elif 'PPP'      in cu: mapa[col] = 'PPP'
    df = df.rename(columns=mapa)

    for req in ['CONTRATO', 'TIPO', 'EMISORA', 'TITULOS', 'IMPORTE_BRUTO', 'PPP']:
        if req not in df.columns:
            raise ValueError(f"Columna requerida no encontrada: {req}")

    df['EMISORA']  = df['EMISORA'].astype(str).str.strip()
    df['CONTRATO'] = df['CONTRATO'].astype(str).str.strip()

    fecha_op = None
    if 'FECHA' in df.columns:
        fechas = df[df['FECHA'].notna()]['FECHA']
        if not fechas.empty:
            f = fechas.iloc[0]
            fecha_op = f if isinstance(f, datetime.datetime) else None

    return df, fecha_op


# ─────────────────────────────────────────────────────────────────────────────
#  CÁLCULO DE PNL — Long, Short e incompletas
# ─────────────────────────────────────────────────────────────────────────────
def calcular_pnl(df):
    totales = df[df['CONTRATO'] == 'TOTAL'].copy()

    compras = (totales[totales['TIPO'] == 'C']
               [['EMISORA', 'TITULOS', 'IMPORTE_BRUTO', 'PPP']].copy()
               .rename(columns={'TITULOS': 'TITULOS_C', 'IMPORTE_BRUTO': 'IMPORTE_C', 'PPP': 'PPP_C'}))
    ventas  = (totales[totales['TIPO'] == 'V']
               [['EMISORA', 'TITULOS', 'IMPORTE_BRUTO', 'PPP']].copy()
               .rename(columns={'TITULOS': 'TITULOS_V', 'IMPORTE_BRUTO': 'IMPORTE_V', 'PPP': 'PPP_V'}))

    pnl = pd.merge(compras, ventas, on='EMISORA', how='outer')

    orden = (totales[totales['TIPO'].isin(['C', 'V'])]
             .groupby(['EMISORA', 'TIPO'])
             .apply(lambda g: g.index.min())
             .unstack('TIPO')
             .rename(columns={'C': 'IDX_C', 'V': 'IDX_V'}))
    pnl = pnl.join(orden, on='EMISORA')

    def clasificar(row):
        tiene_c = pd.notna(row.get('IMPORTE_C'))
        tiene_v = pd.notna(row.get('IMPORTE_V'))
        if tiene_c and tiene_v:
            idx_c = row.get('IDX_C', float('inf'))
            idx_v = row.get('IDX_V', float('inf'))
            return 'LONG' if idx_c <= idx_v else 'SHORT'
        elif tiene_c:
            return 'SOLO_COMPRA'
        elif tiene_v:
            return 'SOLO_VENTA'
        return 'DESCONOCIDO'

    pnl['TIPO_OP'] = pnl.apply(clasificar, axis=1)

    def entrada_ppp(row):
        if row['TIPO_OP'] in ('LONG', 'SOLO_COMPRA'): return row.get('PPP_C')
        if row['TIPO_OP'] in ('SHORT', 'SOLO_VENTA'): return row.get('PPP_V')
        return None

    def salida_ppp(row):
        if row['TIPO_OP'] == 'LONG':  return row.get('PPP_V')
        if row['TIPO_OP'] == 'SHORT': return row.get('PPP_C')
        return None

    def titulos(row):
        if row['TIPO_OP'] in ('LONG', 'SOLO_COMPRA'): return row.get('TITULOS_C')
        if row['TIPO_OP'] in ('SHORT', 'SOLO_VENTA'): return row.get('TITULOS_V')
        return row.get('TITULOS_C') or row.get('TITULOS_V')

    pnl['PPP_ENTRADA'] = pnl.apply(entrada_ppp, axis=1)
    pnl['PPP_SALIDA']  = pnl.apply(salida_ppp, axis=1)
    pnl['TITULOS']     = pnl.apply(titulos, axis=1)

    pnl['PNL'] = pnl['IMPORTE_V'].fillna(0) - pnl['IMPORTE_C'].fillna(0)

    def pnl_pct(row):
        if row['TIPO_OP'] == 'LONG' and pd.notna(row.get('PPP_C')) and row['PPP_C'] != 0:
            return (row['PPP_V'] - row['PPP_C']) / row['PPP_C']
        if row['TIPO_OP'] == 'SHORT' and pd.notna(row.get('PPP_V')) and row['PPP_V'] != 0:
            return (row['PPP_V'] - row['PPP_C']) / row['PPP_V']
        return None

    pnl['PNL_PCT'] = pnl.apply(pnl_pct, axis=1)

    orden_tipo = {'LONG': 0, 'SHORT': 1, 'SOLO_VENTA': 2, 'SOLO_COMPRA': 3, 'DESCONOCIDO': 4}
    pnl['_ORD'] = pnl['TIPO_OP'].map(orden_tipo)
    pnl = pnl.sort_values(['_ORD', 'PNL'], ascending=[True, False]).drop(columns='_ORD').reset_index(drop=True)

    return pnl


# ─────────────────────────────────────────────────────────────────────────────
#  GENERAR EXCEL EN MEMORIA
# ─────────────────────────────────────────────────────────────────────────────
def generar_excel_bytes(pnl, fecha_op):
    def fill(h): return PatternFill("solid", fgColor=h)
    thin  = Side(border_style="thin", color="CCCCCC")
    medio = Side(border_style="medium", color="533483")
    B_CELL  = Border(left=thin, right=thin, top=thin, bottom=thin)
    B_TOTAL = Border(left=thin, right=thin, top=medio, bottom=medio)
    CENTER = Alignment(horizontal="center", vertical="center")
    RIGHT  = Alignment(horizontal="right", vertical="center", indent=1)
    LEFT   = Alignment(horizontal="left", vertical="center", indent=1)
    FMT_CURR, FMT_INT, FMT_PCT = '#,##0.00', '#,##0', '0.00%'

    fecha_str = fecha_op.strftime('%d/%m/%Y') if fecha_op else datetime.date.today().strftime('%d/%m/%Y')
    completas   = pnl[pnl['TIPO_OP'].isin(['LONG', 'SHORT'])]
    incompletas = pnl[pnl['TIPO_OP'].isin(['SOLO_COMPRA', 'SOLO_VENTA'])]
    pnl_total   = completas['PNL'].sum()

    wb = Workbook()
    ws = wb.active
    ws.title = "PNL del Día"
    for col, w in zip("ABCDEFGHI", [18, 10, 14, 14, 17, 17, 15, 11, 14]):
        ws.column_dimensions[col].width = w

    ws.row_dimensions[1].height = 34
    ws.merge_cells("A1:I1")
    c = ws["A1"]; c.value = f"REPORTE PNL DIARIO  —  {fecha_str}"
    c.font = Font(name="Arial", bold=True, color="FFFFFF", size=13); c.fill = fill("1A1A2E"); c.alignment = LEFT

    ws.row_dimensions[2].height = 16
    ws.merge_cells("A2:E2")
    c = ws["A2"]; c.value = "Punto Casa de Bolsa  |  Mesa de Capitales"
    c.font = Font(name="Arial", italic=True, color="AAAAAA", size=8); c.fill = fill("16213E"); c.alignment = LEFT
    ws.merge_cells("F2:I2")
    c = ws["F2"]; c.value = "Cuenta: 9991   |   Fuente: COEBMV101"
    c.font = Font(name="Arial", italic=True, color="AAAAAA", size=8); c.fill = fill("16213E"); c.alignment = LEFT

    ws.row_dimensions[3].height = 20
    hdrs = ["EMISORA", "TÍTULOS", "PPP ENTRADA", "PPP SALIDA", "IMP. COMPRA", "IMP. VENTA", "PNL ($)", "PNL (%)", "TIPO OP."]
    for j, h in enumerate(hdrs, 1):
        c = ws.cell(row=3, column=j, value=h)
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=9); c.fill = fill("533483"); c.alignment = CENTER; c.border = B_CELL

    COLOR_SHORT, COLOR_INC, COLOR_ALT = "FFF3E0", "FFF8E1", "F2EFF7"
    etiquetas = {'LONG': '↑ Long', 'SHORT': '↓ Short', 'SOLO_COMPRA': '⚠ Solo C', 'SOLO_VENTA': '⚠ Solo V'}
    fila = 4
    for i, row in pnl.iterrows():
        ws.row_dimensions[fila].height = 16
        pv, tipo_op = row['PNL'], row['TIPO_OP']
        if tipo_op == 'SHORT': bg = COLOR_SHORT
        elif tipo_op in ('SOLO_COMPRA', 'SOLO_VENTA'): bg = COLOR_INC
        else: bg = "FFFFFF" if i % 2 == 0 else COLOR_ALT
        pc = "1B6B3A" if pv >= 0 else "C0392B"
        celdas = [
            (row['EMISORA'], LEFT, None, Font(name="Arial", bold=True, size=9)),
            (row.get('TITULOS'), RIGHT, FMT_INT, Font(name="Arial", size=9)),
            (row.get('PPP_ENTRADA'), RIGHT, FMT_CURR, Font(name="Arial", size=9)),
            (row.get('PPP_SALIDA'), RIGHT, FMT_CURR, Font(name="Arial", size=9)),
            (row.get('IMPORTE_C') if pd.notna(row.get('IMPORTE_C')) else None, RIGHT, FMT_CURR, Font(name="Arial", size=9)),
            (row.get('IMPORTE_V') if pd.notna(row.get('IMPORTE_V')) else None, RIGHT, FMT_CURR, Font(name="Arial", size=9)),
            (pv if tipo_op not in ('SOLO_COMPRA', 'SOLO_VENTA') else None, RIGHT, FMT_CURR, Font(name="Arial", bold=True, size=9, color=pc)),
            (row.get('PNL_PCT'), RIGHT, FMT_PCT, Font(name="Arial", bold=True, size=9, color=pc)),
            (etiquetas.get(tipo_op, tipo_op), CENTER, None, Font(name="Arial", size=8, italic=True)),
        ]
        for j, (val, align, fmt, fnt) in enumerate(celdas, 1):
            c = ws.cell(row=fila, column=j, value=val)
            c.fill = fill(bg); c.alignment = align; c.border = B_CELL
            if fmt and val is not None: c.number_format = fmt
            if fnt: c.font = fnt
        fila += 1

    if not incompletas.empty:
        ws.row_dimensions[fila].height = 8
        for j in range(1, 10):
            ws.cell(row=fila, column=j, value="").fill = fill("E8E0F0")
        fila += 1

    ws.row_dimensions[fila].height = 22
    tc = "1B6B3A" if pnl_total >= 0 else "C0392B"
    tot = [
        ("TOTAL (ops. completas)", LEFT, None, Font(name="Arial", bold=True, size=9)),
        ("", RIGHT, None, None), ("", RIGHT, None, None), ("", RIGHT, None, None),
        (completas['IMPORTE_C'].sum() if not completas.empty else None, RIGHT, FMT_CURR, Font(name="Arial", bold=True, size=9)),
        (completas['IMPORTE_V'].sum() if not completas.empty else None, RIGHT, FMT_CURR, Font(name="Arial", bold=True, size=9)),
        (pnl_total, RIGHT, FMT_CURR, Font(name="Arial", bold=True, size=11, color=tc)),
        ("", RIGHT, None, None), ("", CENTER, None, None),
    ]
    for j, (val, align, fmt, fnt) in enumerate(tot, 1):
        c = ws.cell(row=fila, column=j, value=val)
        c.fill = fill("EDE7F6"); c.alignment = align; c.border = B_TOTAL
        if fmt and val is not None: c.number_format = fmt
        if fnt: c.font = fnt
    ws.freeze_panes = "A4"

    # Hoja resumen
    ws2 = wb.create_sheet("Resumen")
    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 25
    ws2.row_dimensions[1].height = 30
    ws2.merge_cells("A1:B1")
    c = ws2["A1"]; c.value = f"RESUMEN EJECUTIVO  —  {fecha_str}"
    c.font = Font(name="Arial", bold=True, color="FFFFFF", size=12); c.fill = fill("1A1A2E"); c.alignment = LEFT

    n_long  = int((pnl['TIPO_OP'] == 'LONG').sum())
    n_short = int((pnl['TIPO_OP'] == 'SHORT').sum())
    n_inc   = int(pnl['TIPO_OP'].isin(['SOLO_COMPRA', 'SOLO_VENTA']).sum())
    n_win   = int((completas['PNL'] > 0).sum()) if not completas.empty else 0
    n_loss  = int((completas['PNL'] < 0).sum()) if not completas.empty else 0
    mejor = completas.iloc[0] if not completas.empty else None
    peor  = completas.iloc[-1] if not completas.empty else None
    kpis = [
        ("Emisoras operadas (completas)", len(completas), None, False),
        ("  · Long", n_long, None, False),
        ("  · Short", n_short, None, False),
        ("  · Posiciones incompletas", n_inc, None, False),
        ("Ganadoras / Perdedoras", f"{n_win} / {n_loss}", None, False),
        ("PNL TOTAL ($)", pnl_total, FMT_CURR, True),
        ("Mejor Trade", f"{mejor['EMISORA']}  +${mejor['PNL']:,.2f}" if mejor is not None else "—", None, False),
        ("Peor Trade", f"{peor['EMISORA']}  ${peor['PNL']:,.2f}" if peor is not None else "—", None, False),
        ("Importe Total Comprado", completas['IMPORTE_C'].sum() if not completas.empty else 0, FMT_CURR, False),
        ("Importe Total Vendido", completas['IMPORTE_V'].sum() if not completas.empty else 0, FMT_CURR, False),
    ]
    for i, (label, val, fmt, es_total) in enumerate(kpis, 3):
        ws2.row_dimensions[i].height = 20
        cl = ws2.cell(row=i, column=1, value=label)
        cl.font = Font(name="Arial", bold=not label.startswith("  "), size=9)
        cl.fill = fill("EDE7F6"); cl.alignment = LEFT; cl.border = B_CELL
        cv = ws2.cell(row=i, column=2, value=val)
        cv.alignment = RIGHT; cv.border = B_CELL; cv.fill = fill("FFFFFF")
        if fmt and isinstance(val, (int, float)): cv.number_format = fmt
        cv.font = (Font(name="Arial", bold=True, size=12, color="1B6B3A" if pnl_total >= 0 else "C0392B")
                   if es_total else Font(name="Arial", size=9))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─────────────────────────────────────────────────────────────────────────────
#  TABLA HTML PARA OUTLOOK (se pega con formato en el correo)
# ─────────────────────────────────────────────────────────────────────────────
def construir_tabla_outlook(completas, pnl_total, fecha_str):
    """Devuelve (html, texto_plano) listos para copiar al portapapeles."""
    filas_html = ""
    for i, (_, r) in enumerate(completas.iterrows()):
        bg = "#FFFFFF" if i % 2 == 0 else "#F2EFF7"
        color = "#1B6B3A" if r['PNL'] >= 0 else "#C0392B"
        signo = "+" if r['PNL'] >= 0 else ""
        filas_html += (
            f'<tr style="background:{bg};">'
            f'<td style="border:1px solid #CCCCCC;padding:5px 10px;'
            f'font-family:Arial,sans-serif;font-size:13px;font-weight:bold;">{r["EMISORA"]}</td>'
            f'<td style="border:1px solid #CCCCCC;padding:5px 10px;text-align:right;'
            f'font-family:Arial,sans-serif;font-size:13px;font-weight:bold;color:{color};">'
            f'{signo}${r["PNL"]:,.2f}</td>'
            f'</tr>'
        )

    color_tot = "#1B6B3A" if pnl_total >= 0 else "#C0392B"
    signo_tot = "+" if pnl_total >= 0 else ""
    fila_total = (
        f'<tr style="background:#EDE7F6;">'
        f'<td style="border:2px solid #533483;padding:6px 10px;'
        f'font-family:Arial,sans-serif;font-size:14px;font-weight:bold;">TOTAL</td>'
        f'<td style="border:2px solid #533483;padding:6px 10px;text-align:right;'
        f'font-family:Arial,sans-serif;font-size:14px;font-weight:bold;color:{color_tot};">'
        f'{signo_tot}${pnl_total:,.2f}</td>'
        f'</tr>'
    )

    html = (
        f'<table style="border-collapse:collapse;border:1px solid #CCCCCC;">'
        f'<thead>'
        f'<tr><th colspan="2" style="background:#1A1A2E;color:#FFFFFF;'
        f'padding:8px 10px;text-align:left;font-family:Arial,sans-serif;font-size:13px;">'
        f'PNL Diario &mdash; {fecha_str} &middot; Punto Casa de Bolsa</th></tr>'
        f'<tr>'
        f'<th style="background:#533483;color:#FFFFFF;border:1px solid #CCCCCC;'
        f'padding:5px 10px;text-align:left;font-family:Arial,sans-serif;font-size:12px;">EMISORA</th>'
        f'<th style="background:#533483;color:#FFFFFF;border:1px solid #CCCCCC;'
        f'padding:5px 10px;text-align:right;font-family:Arial,sans-serif;font-size:12px;">PNL</th>'
        f'</tr>'
        f'</thead><tbody>{filas_html}{fila_total}</tbody></table>'
    )

    # Texto plano (tab-separado) como respaldo
    lineas = [f"PNL Diario — {fecha_str} · Punto Casa de Bolsa", "EMISORA\tPNL"]
    for _, r in completas.iterrows():
        signo = "+" if r['PNL'] >= 0 else ""
        lineas.append(f"{r['EMISORA']}\t{signo}${r['PNL']:,.2f}")
    signo_tot = "+" if pnl_total >= 0 else ""
    lineas.append(f"TOTAL\t{signo_tot}${pnl_total:,.2f}")
    texto = "\n".join(lineas)

    return html, texto


# ─────────────────────────────────────────────────────────────────────────────
#  INTERFAZ
# ─────────────────────────────────────────────────────────────────────────────
archivo = st.file_uploader(
    "Sube el archivo COEBMV101 del día",
    type=["xlsx"],
    help="Arrastra el Excel que descargas de tu sistema. Se procesa al instante.",
)

if archivo is not None:
    try:
        file_bytes = archivo.read()
        df, fecha_op = parsear_coebmv(file_bytes)
        pnl = calcular_pnl(df)

        completas = pnl[pnl['TIPO_OP'].isin(['LONG', 'SHORT'])]
        incompletas = pnl[pnl['TIPO_OP'].isin(['SOLO_COMPRA', 'SOLO_VENTA'])]
        pnl_total = completas['PNL'].sum()
        fecha_str = fecha_op.strftime('%d/%m/%Y') if fecha_op else "hoy"

        # ── KPIs ─────────────────────────────────────────────────────────
        st.markdown(f"### Resultados — {fecha_str}")
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("PNL Total", f"${pnl_total:,.2f}",
                  delta=f"{(completas['PNL'] > 0).sum()} ganadoras")
        c2.metric("Emisoras", f"{len(completas)}",
                  delta=f"{int((pnl['TIPO_OP'] == 'SHORT').sum())} shorts")
        if not completas.empty:
            mejor = completas.iloc[0]
            peor  = completas.iloc[-1]
            c3.metric("Mejor trade", mejor['EMISORA'], delta=f"+${mejor['PNL']:,.0f}")
            c4.metric("Peor trade", peor['EMISORA'], delta=f"${peor['PNL']:,.0f}")

        if not incompletas.empty:
            inc_list = ", ".join(incompletas['EMISORA'].tolist())
            st.warning(f"⚠ Posiciones incompletas (revisar manualmente): {inc_list}. "
                       "Aparecen en el reporte pero se excluyen del PNL Total.")

        # ── Tabla en pantalla (solo Emisora y PNL) ───────────────────────
        st.markdown("#### Detalle por emisora")
        disp = pnl[['EMISORA', 'PNL']].copy()
        disp.columns = ['Emisora', 'PNL ($)']

        def color_pnl(v):
            if pd.isna(v):
                return ''
            return 'color: #1B6B3A; font-weight: 700' if v >= 0 else 'color: #C0392B; font-weight: 700'

        styled = (disp.style
                  .map(color_pnl, subset=['PNL ($)'])
                  .format({'PNL ($)': '{:,.2f}'}, na_rep="—"))
        st.dataframe(styled, use_container_width=True, height=min(620, 60 + 35 * len(disp)))

        # ── Copiar para Outlook ──────────────────────────────────────────
        st.markdown("#### Pegar en Outlook")
        tabla_html, tabla_texto = construir_tabla_outlook(completas, pnl_total, fecha_str)
        html_js  = json.dumps(tabla_html)
        text_js  = json.dumps(tabla_texto)

        components.html(f"""
        <div style="font-family:Arial,sans-serif;">
            <button id="btnCopy" onclick="copiarOutlook()"
                style="background:#533483;color:#fff;border:none;border-radius:8px;
                       padding:10px 18px;font-size:14px;font-weight:600;cursor:pointer;
                       width:100%;">
                📋 Copiar tabla para Outlook
            </button>
            <div id="status" style="margin-top:8px;font-size:13px;color:#1B6B3A;height:18px;"></div>
            <div style="margin-top:10px;border:1px dashed #CCC;border-radius:8px;padding:10px;
                        background:#FAFAFA;overflow:auto;">
                <div style="font-size:11px;color:#888;margin-bottom:6px;">
                    Vista previa (también puedes seleccionar y copiar manualmente):
                </div>
                <div id="preview">{tabla_html}</div>
            </div>
        </div>
        <script>
        async function copiarOutlook() {{
            const html = {html_js};
            const text = {text_js};
            const status = document.getElementById('status');
            try {{
                const blobHtml = new Blob([html], {{type: 'text/html'}});
                const blobText = new Blob([text], {{type: 'text/plain'}});
                await navigator.clipboard.write([
                    new ClipboardItem({{'text/html': blobHtml, 'text/plain': blobText}})
                ]);
                status.style.color = '#1B6B3A';
                status.textContent = '✓ Copiado. Pega en Outlook con Ctrl+V (mantiene formato).';
            }} catch (e) {{
                // Respaldo: copiar solo texto plano
                try {{
                    await navigator.clipboard.writeText(text);
                    status.style.color = '#B8860B';
                    status.textContent = '✓ Copiado como texto. (Tu navegador no permitió el formato enriquecido.)';
                }} catch (e2) {{
                    status.style.color = '#C0392B';
                    status.textContent = '✗ No se pudo copiar. Selecciona la tabla de abajo y usa Ctrl+C.';
                }}
            }}
        }}
        </script>
        """, height=180 + 30 * min(len(completas) + 2, 16))

        # ── Descarga ─────────────────────────────────────────────────────
        excel_bytes = generar_excel_bytes(pnl, fecha_op)
        nombre = f"PNL_{fecha_op.strftime('%Y%m%d') if fecha_op else datetime.date.today().strftime('%Y%m%d')}.xlsx"
        st.download_button(
            "⬇  Descargar Excel formateado",
            data=excel_bytes,
            file_name=nombre,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True,
        )

    except Exception as e:
        st.error(f"❌ Error al procesar el archivo: {e}")
        st.info("Verifica que sea el COEBMV101 con las columnas CONTRATO, COMPRA/VENTA, EMISORA, TÍTULOS, IMPORTE BRUTO y PPP.")
else:
    st.info("👆 Sube el COEBMV101 para generar el PNL del día.")
    with st.expander("¿Cómo funciona?"):
        st.markdown("""
        1. **Filtra** las filas `TOTAL` (una por emisora y lado).
        2. **Empareja** Compras (C) y Ventas (V) por emisora.
        3. **Detecta el sentido** de cada operación según cuál apareció primero:
           - **Long** (C → V): compra primero, vende al cierre.
           - **Short** (V → C): vende primero, compra para cerrar.
        4. **Calcula** PNL = Importe Venta − Importe Compra (válido para ambos).
        5. **Marca** posiciones incompletas (solo C o solo V) y las excluye del total.
        6. **Genera** el Excel con detalle por emisora y resumen ejecutivo.
        """)
